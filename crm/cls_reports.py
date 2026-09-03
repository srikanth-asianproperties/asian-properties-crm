"""
=============================================================
cls_reports.py — Asian Properties CRM (APX) | v0.6.1 Reports Enhancements
=============================================================
Version : 1.5
Author  : Built for Asian Properties / Srikanth

WHAT THIS IS
------------
v1.0 shipped 12 reports as a flat card grid, all point-in-time or fixed-
window, plain HTML tables. v1.1 (this version) adds: a date-range picker
on every non-live report (Requirement 1), a categorized landing page
(Requirement 2), 3 new per-salesperson Weekly Reports (Requirement 3), a
new "Lead Stage Analysis" chart report with 4 views (Requirement 4), a
renamed + extended Salesperson Scorecard (Requirement 5, was "Salesperson
Daily Scorecard"), and a new Campaign Insights category with 6 reports
(Requirement 6). 10 new report entries total (22 vs the original 12).

Still owns report metadata, table/column/chart shaping, role-scoping, and
Excel export — still calls into cls_db.py for every database read and
never opens sqlite3 itself.

CAMPAIGN DATA — READ BEFORE TRUSTING THOSE NUMBERS
----------------------------------------------------
leads.campaign is blank/NULL on 99.97% of leads as of 2026-07-17 (see
cls_db.py v2.13's changelog for the exact count) — it's a manual, optional
field, never auto-populated by the Meta/Sell.do pipeline jobs. Srikanth's
explicit call (2026-07): build the Campaign Insights category now anyway,
because he intends to backfill campaign on existing leads. Every campaign-
grouped report therefore buckets blank/NULL under the literal label
"Unknown/Manual" via cls_db._campaign_bucket() rather than dropping those
leads, and right now that one bucket will hold almost everything. This is
expected, not a bug — it'll fill in as campaign gets backfilled/captured
going forward.

CHARTING
--------
Chart.js via CDN (Srikanth's confirmed call, 2026-07) — loaded only by
report_view_charts.html (report_view.html, used by every plain-table
report, gains no new dependency). A report's builder attaches an optional
"chart" key (single-view reports: Campaign Insights C1/C2/C3/C5/C6) or
"views" key (multi-view reports: Lead Stage Analysis's 4 sub-views) to
its returned table dict; build_report() below passes these straight
through. The underlying {columns, rows} table is ALWAYS still produced
alongside every chart — charts are a visual layer on top of the same
accessible/exportable tabular data, never a replacement for it, so Excel
export needs no chart-specific code path (see export_to_excel() / the new
_report_sheets() helper: multi-view reports export one worksheet per
view, everything else exports its one table as before).

DATE RANGE
----------
Every report except the 3 explicitly live/point-in-time ones (Pipeline
Funnel Snapshot, Lead Score Distribution, Owner-wise Workload — "date_
default": "live" in REPORTS below) now takes an optional date range via
?from=YYYY-MM-DD&to=YYYY-MM-DD. Each report carries its own sensible
default ("today" / "this_week" / "this_month" / "last_30_days") in its
REPORTS entry — config-not-code, not inferred from cadence — resolved by
resolve_date_range() below, which app.py's routes call once and thread
into build_report()/export_to_excel() identically, so Excel export always
honors whatever range is currently showing on screen.

Week-start convention: Monday-Sunday (Srikanth's confirmed call, 2026-07)
— see _this_week_range().

ROLE SCOPING (unchanged from v1.0 — same predicate app.py already uses
everywhere else):
  cls_db.can_view_all_leads(role) True  -> admin/manager, sees company-wide
                                            numbers on every owner-filterable
                                            report, and may open admin_only
                                            reports.
  cls_db.can_view_all_leads(role) False -> salesperson, every owner-
                                            filterable report is silently
                                            scoped to their own
                                            owner_match_name, and admin_only
                                            reports 404 for them (app.py
                                            enforces this) rather than just
                                            being hidden from the landing page.

PDF export is still NOT implemented here — browser print-to-PDF, per
Srikanth's v0.6 call (see report_view.html's <style media="print"> block).
The new report_view_charts.html gets the same print treatment, with
canvases hidden and each view's table shown instead — Chart.js canvases
don't print reliably, and the underlying table is the honest fallback.

CHANGELOG
---------
v1.5 (2026-09-02) — Sittings Done, date-range reporting fix. Additive
  only, nothing else touched. Reuses the already-built Salesperson
  Scorecard (date-range picker, Last 7 Days quick-select) rather than
  building anything new — "Sittings Done" previously only existed in
  cls_db.get_todays_activity_counts() (hardcoded to today).
    - _build_salesperson_scorecard()'s columns list: NEW
      ("sittings_done", "Sittings Done") entry, positioned after
      "site_visits_conducted". Data comes for free from cls_db.
      ACTIVITY_METRIC_MAP's new 'sitting_done' -> 'sittings_done' entry
      (cls_db.py this version) — _per_salesperson_activity()'s existing
      cls_db.get_activity_counts_range() call already returns this key
      once that map has it, no other code path change needed here.

v1.4 (2026-08-17) — Monday Weekly Report, new Weekly Reports entry.
  NEW _build_monday_weekly_report() + "monday-weekly-report" REPORTS
  entry, added to the existing Weekly Reports category (after
  "weekend-site-visits"). Reads the SAME cls_db functions the Telegram
  script cls_monday_weekly_report.py uses — get_monday_weekly_report_
  stats() (v2.60, company totals) and get_monday_weekly_report_by_owner()
  (v2.61, per-salesperson) — so numbers here always match what Telegram
  sent for the same week. Two views (Summary + Per Salesperson), same
  "views" shape "weekend-site-visits" already established. Unlike that
  report, this one IS date-range-driven ("date_default": "last_week")
  since week_start/week_end are plain params, not a hardcoded "always
  upcoming weekend" query — resolve_date_range()/default_range_for()
  already handle any QUICK_RANGES key generically, no new plumbing
  needed there.
    - BUG FIX (found while wiring this in, not a pre-existing symptom):
      "last_week" was already a valid REPORT_DATE_PRESETS key (used by
      the picker widget's dropdown, added in v1.2) but was NEVER in the
      separate QUICK_RANGES dict that default_range_for() indexes for a
      report's fallback default — so setting "date_default": "last_week"
      on any REPORTS entry would have raised an uncaught KeyError the
      first time that report was opened with no ?from=&to=/preset query
      param (i.e. every normal click from the landing page). Fixed by
      registering QUICK_RANGES["last_week"] = _last_week_range as a
      statement placed directly after _last_week_range()'s own def
      (not inside the QUICK_RANGES dict literal above, which executes
      before that function exists in the file — doing it there raises
      NameError at import time instead). No other REPORTS entry uses
      "last_week" as its date_default today, so this fix changes no
      existing report's behavior — additive only.
  Non-admin scoping done at the Python layer (filter the by-owner list
  down to current_user["owner_match_name"]) rather than a new SQL owner
  param, since get_monday_weekly_report_by_owner() has none — documented
  in the builder's own docstring. Deliberately NOT reconciled with
  weekly-leads-received/weekly-site-visits-conducted/weekly-site-visits-
  scheduled (different, activity_log actor-attributed definition) —
  Srikanth confirmed 2026-08-17 differing numbers for the same week
  across these reports is expected, not a bug; see this report's
  "caveat" text. ADDITIVE ONLY except the QUICK_RANGES fix described
  above — nothing else existing removed or modified. app.py UNCHANGED —
  the existing generic report_view() route already handles any
  REPORTS-registered id.
v1.3 (2026-08-17) — Weekend Site Visits report, new Weekly Reports entry.
  NEW _build_weekend_site_visits() + "weekend-site-visits" REPORTS entry,
  added to the existing Weekly Reports category. Reads
  cls_db.get_weekend_site_visits() (v2.59) — always "the upcoming
  Saturday+Sunday", never a Srikanth-selectable range, so this uses the
  existing "date_default": "live" exemption (same mechanism as Pipeline
  Funnel Snapshot/Score Distribution/Owner-wise Workload) rather than
  inventing a new no-date-picker path. Two views (Summary + Detail),
  same "views" shape Lead Stage Analysis already established — "columns":
  [], "rows": [], "views": [...] — and "template": "report_view_charts.html"
  since that's the only template that renders report.views (neither view
  carries a "chart" key, which report_view_charts.html already treats as
  optional per view). Same oversight-scoping pattern as every other
  Weekly Reports builder (cls_db.can_view_all_leads(current_user["role"])
  decides company-wide vs current_user["owner_match_name"]).

v1.2 (July 2026) — 3 scoped tasks from Srikanth, all additive.

  TASK 1 — date-range picker dropdown. NEW REPORT_DATE_PRESETS (11 keys,
  same option set as leads_filter.html/app.py's DATE_PRESETS) +
  REPORT_DATE_PRESET_ORDER/REPORT_DATE_PRESET_LABELS +
  _detect_active_report_preset(), deliberately duplicated from app.py
  rather than cross-imported (same convention app.py itself already
  uses). Where a preset key overlaps an existing QUICK_RANGES entry
  (today/this_week/this_month/last_30_days), it reuses that SAME
  function, so "This Week" stays Monday-Sunday (full week) and "This
  Month" stays month-to-date — this file's own existing conventions,
  not app.py's leads-filter variants. resolve_date_range() now also
  accepts an optional `preset` query param, resolved before the
  existing raw from/to validation; a bookmarked ?from=&to= URL with no
  preset still works exactly as before. build_report() now includes
  date_preset/date_preset_order/date_preset_labels in its returned
  dict so _report_date_picker.html needs no new context passed from
  app.py's routes — app.py itself is UNCHANGED. Each report's own
  configured default range (QUICK_RANGES/default_range_for()) is
  UNCHANGED — only the picker widget's option set expanded.

  TASK 2 — mobile transpose. NEW transpose_table(columns, rows,
  label_key) helper. NEW "mobile_transpose": True on 5 REPORTS entries
  (salesperson-scorecard, weekly-site-visits-conducted, weekly-site-
  visits-scheduled, owner-workload, source-performance) — none of
  which use report_view_charts.html. build_report() now additionally
  computes display_columns/display_rows for those 5 (label_key inferred
  as each report's own first column — verified true for all 5), leaving
  columns/rows untouched so export_to_excel()/_report_sheets() need no
  changes.

  TASK 3 — reengagement redefinition (see cls_db.py v2.20). Updated the
  "reengagement" REPORTS entry's caveat text (precise definition, no
  longer "approximate/can't distinguish...") and _build_reengagement()'s
  "Reengaged At" column to read reengaged_at instead of cls_updated_at.

v1.1 (July 2026) — APX v0.6.1 Reports Enhancements. All 10 requirements
  from Srikanth's build prompt, confirmed decisions: Chart.js via CDN,
  build Campaign Insights now (with "Unknown/Manual" fallback bucket) not
  deferred, Monday-Sunday week start, new Scorecard slug
  "salesperson-scorecard" (old "daily-scorecard" 301-redirects, see
  app.py), and all of C1-C6 built including C6 as its own standalone
  report alongside Lead Stage Analysis view D (same underlying
  cls_db.get_site_visits_by_campaign() data, not queried twice).

  NEW date-range helpers: _today_range(), _this_week_range(),
  _this_month_range(), _last_30_days_range(), QUICK_RANGES,
  default_range_for(), resolve_date_range(), quick_select_links().

  NEW STAGE_COLORS + _stage_breakdown_view() — shared by Lead Stage
  Analysis views A/B/C and Campaign Insights C2, reusing the SAME
  cls_db.get_stage_breakdown() call shaped into a {columns, rows, chart}
  dict, so the grouping logic exists in exactly one place.

  NEW _per_salesperson_activity() — shared by the renamed Salesperson
  Scorecard and Weekly Site Visits Conducted/Scheduled (W2/W3), all three
  reading cls_db.get_activity_counts_range() per active salesperson, the
  SAME actor-attributed pattern the v1.0 Daily Scorecard already used
  (see cls_db.py v2.13's changelog on why this is correct instead of a
  raw site_visits/follow_ups.created_by query — that column only ever
  records who SCHEDULED a row, never who conducted/completed it).

  RENAMED "daily-scorecard" -> "salesperson-scorecard", _build_daily_
  scorecard -> _build_salesperson_scorecard, gained 2 columns (Follow-ups
  Created, Site Visits Created) and is now date-range-driven (default
  "today") instead of hardcoded to today. app.py 301-redirects the old
  slug — see that file's changelog.

  NEW report entries (10): weekly-leads-received, weekly-site-visits-
  conducted, weekly-site-visits-scheduled (Weekly Reports category),
  lead-stage-analysis (Pipeline Analysis), campaign-lead-volume,
  campaign-stage-distribution, campaign-conversion-rate, campaign-lost-
  reasons, campaign-response-time, campaign-site-visits (Campaign
  Insights, new Marketing Performance sub-group).

  NEW CATEGORIES + visible_categories() — replaces v1.0's flat
  visible_reports() (single caller, in app.py's reports_home(), updated
  together with this file).

  MODIFIED build_report() — now threads date_from/date_to into every
  builder call and surfaces the new optional "chart"/"views" keys.
  MODIFIED export_to_excel() — generalized to export one worksheet per
  view for multi-view reports via new _report_sheets() helper; single-
  table reports export exactly as v1.0 did.
  MODIFIED 8 existing builders (_build_source_performance, _build_
  project_pipeline, _build_conversion_funnel, _build_hit_rate, _build_
  lost_reason, _build_reengagement, _build_first_response, _build_call_
  activity) — each now accepts date_from/date_to and threads them into
  the cls_db.py v2.13 functions it already called. _build_call_activity
  additionally drops the old daily/weekly period toggle (has_period_
  toggle) in favor of the universal date-range picker — Today is its
  new default, replacing the old toggle's default of weekly.
  UNCHANGED builders (still take **kwargs, ignore date_from/date_to since
  there's no time dimension to apply it to): _build_pipeline_funnel,
  _build_score_distribution, _build_owner_workload.

v1.0 (July 2026) — initial build. All 12 reports from Srikanth's spec.
  Two data-honesty notes surfaced explicitly to the user in the relevant
  reports' `caveat` field (not hidden):
    - Report #5 (Conversion Funnel Over Time): activity_log stage_change
      rows only exist for stage moves made THROUGH the CRM app, not
      Sell.do-driven syncs — see cls_db.py's HONESTY NOTE.
    - Report #3 (Lead Source Performance): built on leads.source
      ('meta'/'selldo_only'/'manual_crm'), not the richer but mostly-NULL
      lead_source_detail — see cls_db.py's HONESTY NOTE.
"""

import io
import re
from datetime import datetime, timedelta

import cls_db

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _scope_owner(current_user):
    """
    The one place that decides "does this login see everyone's numbers
    or just their own" for every owner-filterable report — reuses
    cls_db.can_view_all_leads() exactly as app.py does everywhere else,
    not a new predicate.
    """
    if cls_db.can_view_all_leads(current_user["role"]):
        return None
    return current_user.get("owner_match_name") or "__none__"
    # "__none__" (v1.0): a salesperson login with no owner_match_name set
    # yet (admin hasn't linked their login to a lead_owner value) must
    # see ZERO leads, not everyone's — a value that matches nothing in
    # leads.lead_owner is the deliberate, honest way to enforce that
    # with the existing owner=? scoping, rather than silently falling
    # through to owner=None (which would leak company-wide data to an
    # unlinked login).


# ─────────────────────────────────────────────────────────────
# DATE RANGE — quick-select defaults + query-param resolution
# ─────────────────────────────────────────────────────────────

_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _today_range():
    d = datetime.now().strftime("%Y-%m-%d")
    return d, d


def _this_week_range():
    """Monday-Sunday (Srikanth's confirmed call, 2026-07)."""
    now = datetime.now()
    monday = now - timedelta(days=now.weekday())
    sunday = monday + timedelta(days=6)
    return monday.strftime("%Y-%m-%d"), sunday.strftime("%Y-%m-%d")


def _this_month_range():
    """1st of the current month through TODAY (month-to-date) — not
    through the end of the calendar month, so the range never extends
    into the future."""
    now = datetime.now()
    first = now.replace(day=1)
    return first.strftime("%Y-%m-%d"), now.strftime("%Y-%m-%d")


def _last_30_days_range():
    now = datetime.now()
    start = now - timedelta(days=29)  # inclusive of today = 30 days total
    return start.strftime("%Y-%m-%d"), now.strftime("%Y-%m-%d")


# Config-not-code: every report's "date_default" value (in REPORTS below)
# is a key into this dict. "live" is handled separately (no range at all)
# — see default_range_for().
QUICK_RANGES = {
    "today": _today_range,
    "this_week": _this_week_range,
    "this_month": _this_month_range,
    "last_30_days": _last_30_days_range,
}


def default_range_for(report_id):
    """
    (v1.1) Returns (from_str, to_str) for a report's default range, or
    (None, None) for a "live" (point-in-time, no picker) report.
    """
    meta = REPORTS_BY_ID[report_id]
    kind = meta.get("date_default", "this_month")
    if kind == "live":
        return None, None
    return QUICK_RANGES[kind]()


def resolve_date_range(report_id, args):
    """
    (v1.1) The one place app.py's report routes call to turn a request's
    ?from=&to= query params into the (date_from, date_to) pair every
    builder/export call uses. Falls back to the report's own configured
    default when either param is missing or not a well-formed
    'YYYY-MM-DD' string — a malformed/tampered query string degrades to
    the sensible default rather than erroring.

    args : request.args (or any dict-like with .get())

    Returns (None, None) for a "live" report — callers should treat that
    as "don't show a date picker" (report_view.html/report_view_charts.html
    branch on report.is_live for this).

    v1.2 — Task 1: now ALSO accepts an optional `preset` param (one of
    REPORT_DATE_PRESETS' keys, or "custom"), mirroring app.py's
    _parse_lead_filters() preset-detection logic. A valid non-"custom"
    preset is resolved directly (no _DATE_RE check needed — it comes
    from our own trusted date-math functions, same as app.py's
    DATE_PRESETS[preset]() call). "custom", or no preset at all, falls
    through to the EXISTING raw from/to validation below unchanged — so
    a bookmarked ?from=&to= URL with no preset param still works exactly
    as before; build_report() separately back-fills which preset label
    (if any) those resolved dates match, via _detect_active_report_preset().
    """
    meta = REPORTS_BY_ID.get(report_id)
    if meta is None:
        return None, None
    if meta.get("date_default", "this_month") == "live":
        return None, None
    preset_arg = args.get("preset")
    if preset_arg and preset_arg != "custom" and preset_arg in REPORT_DATE_PRESETS:
        return REPORT_DATE_PRESETS[preset_arg]()
    from_arg = args.get("from")
    to_arg = args.get("to")
    if from_arg and to_arg and _DATE_RE.match(from_arg) and _DATE_RE.match(to_arg) and from_arg <= to_arg:
        return from_arg, to_arg
    return default_range_for(report_id)


def quick_select_links():
    """
    (v1.1) The 4 quick-select buttons every date-range picker shows,
    computed once per request (server-side, so there's no client/server
    timezone math to keep in sync) — [(label, key, from, to), ...].
    "Custom" isn't in this list; it's just the two always-visible
    <input type="date"> fields in the picker partial, pre-filled with
    whatever range is currently showing.

    v1.2: still called by app.py's report_view() route and still a
    valid public helper, but the picker partial itself no longer reads
    its output (see REPORT_DATE_PRESETS below) — left in place rather
    than paused since it's genuinely unchanged and still invoked.
    """
    return [
        ("Today", "today") + _today_range(),
        ("This Week", "this_week") + _this_week_range(),
        ("This Month", "this_month") + _this_month_range(),
        ("Last 30 Days", "last_30_days") + _last_30_days_range(),
    ]


# ─────────────────────────────────────────────────────────────
# DATE RANGE PRESETS (v1.2) — the 11-option picker dropdown, mirroring
# leads_filter.html / app.py's DATE_PRESETS structure (Task 1, July
# 2026). This is a SEPARATE, wider preset list from QUICK_RANGES above
# — QUICK_RANGES/default_range_for() still drive each report's own
# configured DEFAULT range (unchanged by this task); REPORT_DATE_PRESETS
# drives the picker WIDGET's full option set. Deliberately duplicated
# from app.py's DATE_PRESETS rather than cross-imported (same
# duplication convention app.py itself already uses for leads_filter,
# see app.py's changelog on _dr_* — this screen's lifecycle is
# otherwise unrelated to the Leads filter's).
#
# Where a key overlaps an existing QUICK_RANGES entry (today, this_week,
# this_month, last_30_days), it reuses that SAME function — so "This
# Week" here is Monday-Sunday (full week, cls_reports's own existing
# convention, NOT app.py's leads-filter to-date variant) and "This
# Month" is month-to-date, exactly as documented at the top of this
# file. The remaining keys are new, same date math as app.py's _dr_*.
# ─────────────────────────────────────────────────────────────

def _yesterday_range():
    d = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    return d, d


def _last_n_days_range(n):
    now = datetime.now()
    start = now - timedelta(days=n - 1)  # inclusive of today
    return start.strftime("%Y-%m-%d"), now.strftime("%Y-%m-%d")


def _last_week_range():
    """Full previous Monday-Sunday (7 days), not to-date."""
    now = datetime.now()
    this_monday = now - timedelta(days=now.weekday())
    last_monday = this_monday - timedelta(days=7)
    last_sunday = last_monday + timedelta(days=6)
    return last_monday.strftime("%Y-%m-%d"), last_sunday.strftime("%Y-%m-%d")


# v1.4: registered into QUICK_RANGES here (not in the dict literal above,
# which executes before this function is defined) so "date_default":
# "last_week" (Monday Weekly Report) resolves correctly via
# default_range_for() the same way every other QUICK_RANGES key does.
QUICK_RANGES["last_week"] = _last_week_range


def _last_month_range():
    """Full previous calendar month, 1st through last day."""
    now = datetime.now()
    first_this_month = now.replace(day=1)
    last_day_prev_month = first_this_month - timedelta(days=1)
    first_prev_month = last_day_prev_month.replace(day=1)
    return first_prev_month.strftime("%Y-%m-%d"), last_day_prev_month.strftime("%Y-%m-%d")


def _maximum_range():
    """"Maximum" clears the range entirely (all-time), returned as
    ("", "") so it round-trips through the same "" == no-filter
    convention every builder's `if date_from and date_to:` guard
    already uses (see e.g. get_source_performance())."""
    return "", ""


# Config-not-code: the dropdown's option order AND the resolver share
# this one dict — add a preset here and both follow automatically.
REPORT_DATE_PRESETS = {
    "today":        _today_range,
    "yesterday":    _yesterday_range,
    "last_7_days":  lambda: _last_n_days_range(7),
    "last_14_days": lambda: _last_n_days_range(14),
    "last_30_days": _last_30_days_range,
    "this_week":    _this_week_range,
    "last_week":    _last_week_range,
    "this_month":   _this_month_range,
    "last_month":   _last_month_range,
    "maximum":      _maximum_range,
}

REPORT_DATE_PRESET_ORDER = list(REPORT_DATE_PRESETS.keys()) + ["custom"]

REPORT_DATE_PRESET_LABELS = {
    "today": "Today", "yesterday": "Yesterday",
    "last_7_days": "Last 7 days", "last_14_days": "Last 14 days",
    "last_30_days": "Last 30 days",
    "this_week": "This week", "last_week": "Last week",
    "this_month": "This month", "last_month": "Last month",
    "maximum": "Maximum", "custom": "Custom",
}


def _detect_active_report_preset(date_from, date_to):
    """
    Given the currently-resolved date_from/date_to, figure out which
    preset (if any) produced them — used to pre-select the right
    dropdown option, both for a preset-driven navigation and for a
    bookmarked ?from=&to= URL carrying no preset param at all.

    Unlike leads_filter's blank-date convention (blank means "no filter
    active"), every cls_reports report ALWAYS resolves to a concrete
    range — resolve_date_range() never returns ("", "") except via an
    explicit "maximum" selection, so blank/blank is unambiguous here
    and maps directly to "maximum" rather than to an unset state.
    """
    if not date_from and not date_to:
        return "maximum"
    for key, fn in REPORT_DATE_PRESETS.items():
        if key == "maximum":
            continue  # maximum's own output is ("",""), already handled above
        f, t = fn()
        if f == date_from and t == date_to:
            return key
    return "custom"


# ─────────────────────────────────────────────────────────────
# SHARED CHART/GROUPING HELPERS
# ─────────────────────────────────────────────────────────────

# Reuses the exact stage colors base.html's .stage-* badge classes
# already use, so a chart bar and a stage badge elsewhere in the app
# read as the same color for the same stage. Lost gets a distinct
# darker red from Unqualified for chart legibility — base.html's CSS
# intentionally combines them into one badge color, but a bar-chart
# legend needs the two stages visually distinguishable.
STAGE_COLORS = {
    "Incoming": "#8A93A3", "Prospect": "#2F80ED", "Opportunity": "#E08C2B",
    "Site Visited": "#27AE60", "Unqualified": "#D64545", "Lost": "#A93226",
    "Re Assigned": "#7E57C2", "Booked": "#1E8449",
}


def _stage_breakdown_view(data, group_label):
    """
    (v1.1) Shapes cls_db.get_stage_breakdown()'s {group: {stage: count,
    total}} return into a {columns, rows, chart} view dict — one stacked-
    bar dataset per stage, groups sorted by total desc. Shared by Lead
    Stage Analysis views A/B/C and Campaign Insights C2 (Campaign Stage
    Distribution is exactly view C, reused rather than re-queried).
    """
    groups = sorted(data.items(), key=lambda kv: -kv[1]["total"])
    labels = [g for g, _ in groups]
    datasets = [
        {"label": stage, "color": STAGE_COLORS[stage], "data": [data[g][stage] for g in labels]}
        for stage in cls_db.ALL_STAGES
    ]
    columns = [("group", group_label)] + [(s, s) for s in cls_db.ALL_STAGES] + [("total", "Total")]
    rows = [{"group": g, **data[g]} for g in labels]
    return {"columns": columns, "rows": rows, "chart": {"type": "stacked_bar", "labels": labels, "datasets": datasets}}


def _per_salesperson_activity(current_user, date_from, date_to):
    """
    (v1.1) One row per salesperson in scope — every active user for an
    oversight login (admin/manager), just the current login for a
    salesperson — each paired with their cls_db.get_activity_counts_range()
    dict for [date_from, date_to]. Shared by the Salesperson Scorecard and
    Weekly Site Visits Conducted/Scheduled (W2/W3): same per-user loop the
    v1.0 Daily Scorecard already used, just reading the range-aware
    function instead of the today-only one.

    Returns a list of (display_name, activity_counts_dict) tuples.
    """
    rows = []
    if cls_db.can_view_all_leads(current_user["role"]):
        users = [u for u in cls_db.get_all_users_detailed() if u["active"]]
        for u in users:
            perf = cls_db.get_activity_counts_range(actor_email=u["email"], date_from=date_from, date_to=date_to)
            rows.append((u["full_name"] or u["email"], perf))
    else:
        perf = cls_db.get_activity_counts_range(actor_email=current_user["email"], date_from=date_from, date_to=date_to)
        rows.append((current_user["full_name"] or current_user["email"], perf))
    return rows


def transpose_table(columns, rows, label_key):
    """
    (v1.2, Task 2) Swaps a {columns, rows} table's axes for mobile
    display: the original row-identity column (label_key) becomes the
    new header row, and each original metric column becomes a new row.
    For reports shaped one-row-per-person with many metric columns
    (unreadable on a phone-width table), the transposed view reads as
    metric-rows x person-columns instead.

    label_key : the column key whose per-row value becomes each new
                column's header (e.g. "name" or "owner") — excluded
                from the transposed metric rows, since it's now the
                header rather than a metric.

    Returns (display_columns, display_rows), same {key, label} tuple /
    dict-per-row shape report_view.html already knows how to render.
    """
    display_columns = [("metric", "Metric")] + [
        (str(row.get(label_key, "")), row.get(label_key, "")) for row in rows
    ]
    display_rows = []
    for key, label in columns:
        if key == label_key:
            continue
        display_row = {"metric": label}
        for row in rows:
            display_row[str(row.get(label_key, ""))] = row.get(key)
        display_rows.append(display_row)
    return display_columns, display_rows


# ─────────────────────────────────────────────────────────────
# PER-REPORT BUILDERS — each returns {"columns", "rows"} plus an
# optional "chart" (single view) or "views" (multi-view) key.
# ─────────────────────────────────────────────────────────────

def _build_salesperson_scorecard(current_user, date_from=None, date_to=None, **kwargs):
    columns = [
        ("name", "Salesperson"), ("calls_attempted", "Calls Tapped"),
        ("notes_added", "Notes Added"),
        ("follow_ups_created", "Follow-ups Created"),
        ("follow_ups_completed", "Follow-ups Conducted"),
        ("site_visits_created", "Site Visits Created"),
        ("site_visits_conducted", "Site Visits Conducted"),
        ("sittings_done", "Sittings Done"),
    ]
    rows = [{"name": name, **perf} for name, perf in _per_salesperson_activity(current_user, date_from, date_to)]
    if cls_db.can_view_all_leads(current_user["role"]):
        company = cls_db.get_activity_counts_range(actor_email=None, date_from=date_from, date_to=date_to)
        rows.append({"name": "— Company Wide —", **company})
    return {"columns": columns, "rows": rows}


def _build_pipeline_funnel(current_user, **kwargs):
    owner = _scope_owner(current_user)
    counts = cls_db.get_stage_snapshot_counts(owner=owner)
    columns = [("stage", "Stage"), ("count", "Leads")]
    rows = [{"stage": stage, "count": counts[stage]} for stage in cls_db.ALL_STAGES]
    return {"columns": columns, "rows": rows}


def _build_source_performance(current_user, date_from=None, date_to=None, **kwargs):
    data = cls_db.get_source_performance(date_from=date_from, date_to=date_to)
    columns = [
        ("source", "Source"), ("total", "Total Leads"),
        ("prospect_plus", "Reached Prospect+"), ("opportunity_plus", "Reached Opportunity+"),
        ("site_visited_plus", "Reached Site Visited+"), ("booked", "Booked"),
        ("lost_unqualified", "Lost/Unqualified"), ("quality_rate", "Opportunity+ Rate"),
    ]
    rows = []
    for r in data:
        row = dict(r)
        row["quality_rate"] = f'{round(r["opportunity_plus"] / r["total"] * 100, 1)}%' if r["total"] else "—"
        rows.append(row)
    return {"columns": columns, "rows": rows}


def _build_project_pipeline(current_user, date_from=None, date_to=None, **kwargs):
    owner = _scope_owner(current_user)
    data = cls_db.get_project_pipeline(owner=owner, date_from=date_from, date_to=date_to)
    columns = [("project", "Project")] + [(s, s) for s in cls_db.ALL_STAGES] + [("total", "Total")]
    rows = []
    for project, stage_counts in sorted(data.items(), key=lambda kv: -kv[1]["total"]):
        rows.append({"project": project, **stage_counts})
    return {"columns": columns, "rows": rows}


def _build_conversion_funnel(current_user, date_from=None, date_to=None, **kwargs):
    data = cls_db.get_conversion_funnel_trend(date_from=date_from, date_to=date_to)
    columns = [
        ("month", "Month"), ("entered_prospect", "Entered Prospect"),
        ("reached_opportunity", "Reached Opportunity"), ("opportunity_rate", "Opportunity Rate"),
        ("reached_site_visited", "Reached Site Visited"), ("site_visited_rate", "Site Visit Rate"),
    ]
    rows = []
    for r in data:
        row = dict(r)
        row["opportunity_rate"] = f'{r["opportunity_rate"]}%'
        row["site_visited_rate"] = f'{r["site_visited_rate"]}%'
        rows.append(row)
    return {"columns": columns, "rows": rows}


def _build_hit_rate(current_user, date_from=None, date_to=None, **kwargs):
    owner = _scope_owner(current_user)
    data = cls_db.get_followup_hit_rate(owner=owner, date_from=date_from, date_to=date_to)
    columns = [
        ("kind", "Type"), ("completed", "Completed"), ("missed", "Missed"),
        ("upcoming", "Upcoming"), ("cancelled", "Cancelled"), ("no_show", "No-Show"),
        ("total", "Total"),
    ]
    rows = [
        {"kind": "Site Visits", **data["site_visits"]},
        {"kind": "Follow-ups", "no_show": "—", **data["follow_ups"]},
    ]
    return {"columns": columns, "rows": rows}


def _build_lost_reason(current_user, date_from=None, date_to=None, **kwargs):
    data = cls_db.get_lost_reason_breakdown(date_from=date_from, date_to=date_to)
    columns = [("stage", "Stage"), ("reason", "Reason"), ("count", "Count")]
    return {"columns": columns, "rows": data}


def _build_score_distribution(current_user, **kwargs):
    owner = _scope_owner(current_user)
    data = cls_db.get_score_distribution(owner=owner)
    columns = [("owner", "Salesperson"), ("Hot", "Hot"), ("Warm", "Warm"), ("Cold", "Cold"), ("total", "Total")]
    return {"columns": columns, "rows": data}


def _build_reengagement(current_user, date_from=None, date_to=None, **kwargs):
    owner = _scope_owner(current_user)
    leads = cls_db.get_reengaged_leads(owner=owner, date_from=date_from, date_to=date_to)
    columns = [
        ("crm_lead_no", "Lead #"), ("full_name", "Name"), ("current_stage", "Stage"),
        ("lead_owner", "Owner"), ("cls_created_at", "First Created"), ("reengaged_at", "Reengaged At"),
    ]
    rows = [{k: lead.get(k) for k, _ in columns} for lead in leads]
    return {"columns": columns, "rows": rows}


def _build_first_response(current_user, date_from=None, date_to=None, **kwargs):
    data = cls_db.get_first_response_time(date_from=date_from, date_to=date_to)
    columns = [
        ("month", "Month"), ("leads", "Leads"), ("responded", "Responded"),
        ("no_response_yet", "No Response Yet"), ("avg_hours", "Avg Response (hrs)"),
        ("median_hours", "Median Response (hrs)"),
    ]
    rows = []
    for r in data:
        row = dict(r)
        row["avg_hours"] = row["avg_hours"] if row["avg_hours"] is not None else "—"
        row["median_hours"] = row["median_hours"] if row["median_hours"] is not None else "—"
        rows.append(row)
    return {"columns": columns, "rows": rows}


def _build_owner_workload(current_user, **kwargs):
    owner = _scope_owner(current_user)
    data = cls_db.get_owner_workload(owner=owner)
    columns = [
        ("owner", "Salesperson"), ("open_leads", "Open Leads"),
        ("open_follow_ups", "Open Follow-ups"), ("open_site_visits", "Open Site Visits"),
    ]
    return {"columns": columns, "rows": data}


def _build_call_activity(current_user, date_from=None, date_to=None, **kwargs):
    owner = _scope_owner(current_user)
    data = cls_db.get_call_activity(owner=owner, date_from=date_from, date_to=date_to)
    columns = [
        ("crm_lead_no", "Lead #"), ("full_name", "Name"), ("lead_owner", "Owner"),
        ("call_count", "Calls"), ("last_call_at", "Last Call"),
    ]
    return {"columns": columns, "rows": data}


# ── Weekly Reports (Requirement 3) ──

def _build_weekly_leads_received(current_user, date_from=None, date_to=None, **kwargs):
    owner = _scope_owner(current_user)
    data = cls_db.get_leads_received_by_owner(date_from, date_to, owner=owner)
    columns = [("owner", "Salesperson"), ("count", "Leads Received")]
    return {"columns": columns, "rows": data}


def _build_weekly_site_visits_conducted(current_user, date_from=None, date_to=None, **kwargs):
    columns = [("name", "Salesperson"), ("count", "Site Visits Conducted")]
    rows = [{"name": name, "count": perf["site_visits_conducted"]}
            for name, perf in _per_salesperson_activity(current_user, date_from, date_to)]
    return {"columns": columns, "rows": rows}


def _build_weekly_site_visits_scheduled(current_user, date_from=None, date_to=None, **kwargs):
    columns = [("name", "Salesperson"), ("count", "Site Visits Scheduled")]
    rows = [{"name": name, "count": perf["site_visits_created"]}
            for name, perf in _per_salesperson_activity(current_user, date_from, date_to)]
    return {"columns": columns, "rows": rows}


def _format_scheduled_at(value):
    """
    (v1.3) Same 12-hour display convention as app.py's "ampm" Jinja
    filter (which report_view_charts.html's generic table doesn't
    apply per-column) — reproduced here rather than imported to avoid
    a cls_reports -> app import (app.py already imports cls_reports,
    not the other way around). Only site_visits.scheduled_at's
    "YYYY-MM-DD HH:MM" shape is relevant here, but both formats
    ampm_filter() handles are kept so this stays a drop-in match.
    """
    if not value:
        return value
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(value, fmt).strftime("%a, %b %d, %I:%M %p")
        except ValueError:
            continue
    return value


def _build_weekend_site_visits(current_user, **kwargs):
    oversight = cls_db.can_view_all_leads(current_user["role"])
    visits = cls_db.get_weekend_site_visits(
        owner_match_name=None if oversight else current_user["owner_match_name"])

    counts = {}
    for v in visits:
        counts[v["lead_owner"]] = counts.get(v["lead_owner"], 0) + 1
    summary_view = {
        "columns": [("owner", "Salesperson"), ("count", "Visits This Weekend")],
        "rows": [{"owner": owner, "count": count} for owner, count in counts.items()],
    }

    detail_view = {
        "columns": [
            ("owner", "Salesperson"), ("name", "Lead Name"), ("phone", "Phone"),
            ("project", "Project"), ("scheduled", "Scheduled"),
        ],
        "rows": [
            {
                "owner": v["lead_owner"], "name": v["full_name"], "phone": v["phone_raw"],
                "project": v["project"], "scheduled": _format_scheduled_at(v["scheduled_at"]),
            }
            for v in visits
        ],
    }

    views = [
        {"title": "Summary", **summary_view},
        {"title": "Detail", **detail_view},
    ]
    return {"columns": [], "rows": [], "views": views}


def _build_monday_weekly_report(current_user, date_from=None, date_to=None, **kwargs):
    """
    (v1.4) "Monday Weekly Report" — Weekly Reports entry. Reads
    cls_db.get_monday_weekly_report_stats() (company totals) and
    get_monday_weekly_report_by_owner() (per-salesperson), the SAME
    functions cls_monday_weekly_report.py (the Telegram script) uses —
    numbers here will always match what Telegram sent for the same
    week. Deliberately NOT reconciled with the existing weekly-leads-
    received/weekly-site-visits-conducted/weekly-site-visits-scheduled
    reports, which use a different (activity_log actor-attributed)
    definition — Srikanth confirmed 2026-08-17 this may show different
    numbers than those 3 reports for the same week, and that's expected,
    not a bug.

    Non-admin: rows filtered to the logged-in user's own
    owner_match_name via the by-owner breakdown (no separate scoped
    query needed) — same "owner_filterable" convention as every other
    report here, just implemented at the Python layer instead of SQL
    since get_monday_weekly_report_by_owner() has no owner param.
    Admin (cls_db.can_view_all_leads): full company summary + every
    salesperson's row, unfiltered.
    """
    by_owner = cls_db.get_monday_weekly_report_by_owner(date_from, date_to)

    if not cls_db.can_view_all_leads(current_user["role"]):
        name = current_user.get("owner_match_name")
        by_owner = [r for r in by_owner if r["owner"] == name]
        own = by_owner[0] if by_owner else {
            "leads_generated": 0, "site_visits_scheduled": 0,
            "site_visits_conducted": 0, "bookings_total": 0,
        }
        summary_rows = [
            {"metric": "Leads Generated", "value": own["leads_generated"]},
            {"metric": "Site Visits Scheduled", "value": own["site_visits_scheduled"]},
            {"metric": "Site Visits Conducted", "value": own["site_visits_conducted"]},
            {"metric": "Bookings (Total)", "value": own["bookings_total"]},
        ]
    else:
        stats = cls_db.get_monday_weekly_report_stats(date_from, date_to)
        summary_rows = [
            {"metric": "Leads Generated", "value": stats["leads_generated"]},
            {"metric": "Site Visits Scheduled", "value": stats["site_visits_scheduled"]},
            {"metric": "Site Visits Conducted", "value": stats["site_visits_conducted"]},
            {"metric": "Bookings — Weekdays", "value": stats["bookings_weekday"]},
            {"metric": "Bookings — Weekend", "value": stats["bookings_weekend"]},
            {"metric": "Total Bookings", "value": stats["bookings_total"]},
        ]

    views = [
        {
            "title": "Summary",
            "columns": [("metric", "Metric"), ("value", "Value")],
            "rows": summary_rows,
        },
        {
            "title": "Per Salesperson",
            "columns": [
                ("owner", "Salesperson"), ("leads_generated", "Leads"),
                ("site_visits_scheduled", "Scheduled"),
                ("site_visits_conducted", "Conducted"),
                ("bookings_total", "Bookings"),
            ],
            "rows": by_owner,
        },
    ]
    return {"columns": [], "rows": [], "views": views}


# ── Lead Stage Analysis (Requirement 4) ──

def _build_lead_stage_analysis(current_user, date_from=None, date_to=None, **kwargs):
    owner = _scope_owner(current_user)
    by_owner = _stage_breakdown_view(
        cls_db.get_stage_breakdown("owner", date_from=date_from, date_to=date_to, owner=owner), "Owner")
    by_project = _stage_breakdown_view(
        cls_db.get_stage_breakdown("project", date_from=date_from, date_to=date_to, owner=owner), "Project")
    by_campaign = _stage_breakdown_view(
        cls_db.get_stage_breakdown("campaign", date_from=date_from, date_to=date_to, owner=owner), "Campaign")

    visits = cls_db.get_site_visits_by_campaign(date_from=date_from, date_to=date_to, owner=owner)
    visits_view = {
        "columns": [("campaign", "Campaign"), ("count", "Site Visits")], "rows": visits,
        "chart": {"type": "bar", "labels": [r["campaign"] for r in visits],
                  "datasets": [{"label": "Site Visits", "color": "#2F80ED", "data": [r["count"] for r in visits]}]},
    }

    views = [
        {"title": "By Owner", **by_owner},
        {"title": "By Project", **by_project},
        {"title": "By Campaign", **by_campaign},
        {"title": "Site Visits by Campaign", **visits_view},
    ]
    return {"columns": [], "rows": [], "views": views}


# ── Campaign Insights (Requirement 6) ──

def _build_campaign_lead_volume(current_user, date_from=None, date_to=None, **kwargs):
    data = cls_db.get_campaign_lead_volume(date_from=date_from, date_to=date_to)
    columns = [("campaign", "Campaign"), ("count", "Leads")]
    chart = {"type": "horizontal_bar", "labels": [r["campaign"] for r in data],
             "datasets": [{"label": "Leads", "color": "#2F80ED", "data": [r["count"] for r in data]}]}
    return {"columns": columns, "rows": data, "chart": chart}


def _build_campaign_stage_distribution(current_user, date_from=None, date_to=None, **kwargs):
    data = cls_db.get_stage_breakdown("campaign", date_from=date_from, date_to=date_to)
    return _stage_breakdown_view(data, "Campaign")


def _build_campaign_conversion_rate(current_user, date_from=None, date_to=None, **kwargs):
    data = cls_db.get_campaign_performance(date_from=date_from, date_to=date_to)
    columns = [
        ("campaign", "Campaign"), ("total", "Total Leads"),
        ("prospect_plus", "Reached Prospect+"), ("opportunity_plus", "Reached Opportunity+"),
        ("site_visited_plus", "Reached Site Visited+"), ("booked", "Booked"),
        ("lost_unqualified", "Lost/Unqualified"), ("conversion_rate", "Site Visited+ Rate"),
    ]
    rows = []
    for r in data:
        row = dict(r)
        row["conversion_rate"] = round(r["site_visited_plus"] / r["total"] * 100, 1) if r["total"] else 0.0
        rows.append(row)
    chart = {"type": "bar", "labels": [r["campaign"] for r in rows],
             "datasets": [{"label": "Site Visited+ Rate (%)", "color": "#27AE60",
                           "data": [r["conversion_rate"] for r in rows]}]}
    for row in rows:
        row["conversion_rate"] = f'{row["conversion_rate"]}%'
    return {"columns": columns, "rows": rows, "chart": chart}


def _build_campaign_lost_reasons(current_user, date_from=None, date_to=None, **kwargs):
    data = cls_db.get_campaign_lost_reasons(date_from=date_from, date_to=date_to)
    columns = [("campaign", "Campaign"), ("reason", "Reason"), ("count", "Count")]
    return {"columns": columns, "rows": data}


def _build_campaign_response_time(current_user, date_from=None, date_to=None, **kwargs):
    data = cls_db.get_campaign_response_time(date_from=date_from, date_to=date_to)
    columns = [
        ("campaign", "Campaign"), ("leads", "Leads"), ("responded", "Responded"),
        ("no_response_yet", "No Response Yet"), ("avg_hours", "Avg Response (hrs)"),
        ("median_hours", "Median Response (hrs)"),
    ]
    chart = {"type": "bar", "labels": [r["campaign"] for r in data],
             "datasets": [{"label": "Median Response (hrs)", "color": "#E08C2B",
                           "data": [r["median_hours"] or 0 for r in data]}]}
    rows = []
    for r in data:
        row = dict(r)
        row["avg_hours"] = row["avg_hours"] if row["avg_hours"] is not None else "—"
        row["median_hours"] = row["median_hours"] if row["median_hours"] is not None else "—"
        rows.append(row)
    return {"columns": columns, "rows": rows, "chart": chart}


def _build_campaign_site_visits(current_user, date_from=None, date_to=None, **kwargs):
    data = cls_db.get_site_visits_by_campaign(date_from=date_from, date_to=date_to)
    columns = [("campaign", "Campaign"), ("count", "Site Visits")]
    chart = {"type": "bar", "labels": [r["campaign"] for r in data],
             "datasets": [{"label": "Site Visits", "color": "#2F80ED", "data": [r["count"] for r in data]}]}
    return {"columns": columns, "rows": data, "chart": chart}


# ─────────────────────────────────────────────────────────────
# REGISTRY — config-not-code. Add a report by adding ONE entry here;
# every route, the landing page, and the Excel exporter all read from
# this single list, nothing else needs to change. Add it to a category
# in CATEGORIES below too, or it simply won't appear on the landing page
# (build_report()/the route still work by direct URL either way).
#
# date_default: "today" | "this_week" | "this_month" | "last_30_days" |
#               "live" (no date picker — see Requirement 1's 3 exempt
#               reports: pipeline-funnel, score-distribution, owner-workload)
# template: report_view.html (default, omitted) | report_view_charts.html
#           (chart-bearing reports)
# ─────────────────────────────────────────────────────────────

REPORTS = [
    {
        "id": "salesperson-scorecard", "title": "Salesperson Scorecard",
        "description": "Calls tapped, notes added, follow-ups and site visits created/conducted, for the selected period.",
        "cadence": "Daily", "admin_only": False, "owner_filterable": True,
        "caveat": None, "build": _build_salesperson_scorecard, "date_default": "today",
        "mobile_transpose": True,
    },
    {
        "id": "pipeline-funnel", "title": "Pipeline Funnel Snapshot",
        "description": "Count of leads in each of the 8 stages, right now.",
        "cadence": "Live", "admin_only": False, "owner_filterable": True,
        "caveat": "A live snapshot of where leads sit right now, not a historical figure for today specifically.",
        "build": _build_pipeline_funnel, "date_default": "live",
    },
    {
        "id": "source-performance", "title": "Lead Source Performance",
        "description": "Which capture channel (Meta / Sell.do-only / manually entered) produces the best-quality leads.",
        "cadence": "Weekly", "admin_only": True, "owner_filterable": False,
        "caveat": (
            "Based on leads.source (meta / selldo_only / manual_crm) — the system-level capture channel, "
            "not a marketing sub-source like \"99acres\". The richer lead_source_detail field is set only "
            "for manually-entered leads and is NULL for the large majority of leads, so it isn't usable "
            "here yet."
        ),
        "build": _build_source_performance, "date_default": "last_30_days",
        "mobile_transpose": True,
    },
    {
        "id": "project-pipeline", "title": "Project-wise Pipeline",
        "description": "Stage breakdown per project (Naishka Homes / Grace Classic / Prima Paradiso / Praga Enclave), for leads created in the selected period.",
        "cadence": "Weekly", "admin_only": False, "owner_filterable": True,
        "caveat": None, "build": _build_project_pipeline, "date_default": "this_month",
    },
    {
        "id": "conversion-funnel", "title": "Conversion Funnel Over Time",
        "description": "Prospect → Opportunity → Site Visit conversion rate trend, month over month.",
        "cadence": "Monthly", "admin_only": True, "owner_filterable": False,
        "caveat": (
            "Only counts stage moves made THROUGH the CRM app (activity_log 'stage_change' rows) — Sell.do-"
            "driven syncs during parallel-run leave no such trail, so months before the team's CRM adoption "
            "matured will undercount. This is a real data gap, not a bug; it fills in as parallel-run continues."
        ),
        "build": _build_conversion_funnel, "date_default": "this_month",
    },
    {
        "id": "hit-rate", "title": "Follow-up / Site-Visit Hit Rate",
        "description": "Of follow-ups and site visits scheduled in the selected period: completed vs missed vs cancelled.",
        "cadence": "Weekly", "admin_only": False, "owner_filterable": True,
        "caveat": "\"Missed\" is computed live (still scheduled and past due), never a stored status.",
        "build": _build_hit_rate, "date_default": "this_week",
    },
    {
        "id": "lost-reason", "title": "Lost/Unqualified Reason Breakdown",
        "description": "Why leads are dying — price, location, budget, no response.",
        "cadence": "Monthly", "admin_only": True, "owner_filterable": False,
        "caveat": "Reflects each lead's CURRENT reason only — cleared the moment a lead moves out of Lost/Unqualified.",
        "build": _build_lost_reason, "date_default": "this_month",
    },
    {
        "id": "score-distribution", "title": "Lead Score Distribution",
        "description": "Hot / Warm / Cold count, by salesperson.",
        "cadence": "Live", "admin_only": False, "owner_filterable": True,
        "caveat": None, "build": _build_score_distribution, "date_default": "live",
    },
    {
        "id": "reengagement", "title": "Reengagement Report",
        "description": "Which existing leads came back during the selected period.",
        "cadence": "Weekly", "admin_only": False, "owner_filterable": True,
        "caveat": (
            "Precise signal (v2.20) — an existing contact re-entered through Meta (matched by phone/email to a "
            "prior lead) and no activity has been logged against it since. Clears the moment any activity is "
            "logged, computed live, same as \"missed\" statuses elsewhere in the app."
        ),
        "build": _build_reengagement, "date_default": "this_week",
    },
    {
        "id": "first-response", "title": "First-Response Time",
        "description": "Time from lead creation to first activity_log entry.",
        "cadence": "Monthly", "admin_only": True, "owner_filterable": False,
        "caveat": None, "build": _build_first_response, "date_default": "this_month",
    },
    {
        "id": "owner-workload", "title": "Owner-wise Workload",
        "description": "Open leads and open follow-ups/site-visits per salesperson, right now.",
        "cadence": "Live", "admin_only": False, "owner_filterable": True,
        "caveat": None, "build": _build_owner_workload, "date_default": "live",
        "mobile_transpose": True,
    },
    {
        "id": "call-activity", "title": "Call Activity Report",
        "description": "Calls tapped per lead per salesperson, for the selected period (best-effort — no duration; that's v1.0 Telephony).",
        "cadence": "Daily", "admin_only": False, "owner_filterable": True,
        "caveat": None, "build": _build_call_activity, "date_default": "today",
    },
    {
        "id": "weekly-leads-received", "title": "Weekly Leads Received",
        "description": "New leads assigned to each salesperson during the selected week.",
        "cadence": "Weekly", "admin_only": False, "owner_filterable": True,
        "caveat": (
            "Grouped by each lead's CURRENT owner, not who it was assigned to at creation time — a lead "
            "reassigned since shows under its new owner."
        ),
        "build": _build_weekly_leads_received, "date_default": "this_week",
    },
    {
        "id": "weekly-site-visits-conducted", "title": "Weekly Site Visits Conducted",
        "description": "Site visits marked completed by each salesperson during the selected week.",
        "cadence": "Weekly", "admin_only": False, "owner_filterable": True,
        "caveat": None, "build": _build_weekly_site_visits_conducted, "date_default": "this_week",
        "mobile_transpose": True,
    },
    {
        "id": "weekly-site-visits-scheduled", "title": "Weekly Site Visits Scheduled",
        "description": "Site visits created (scheduled) by each salesperson during the selected week.",
        "cadence": "Weekly", "admin_only": False, "owner_filterable": True,
        "caveat": None, "build": _build_weekly_site_visits_scheduled, "date_default": "this_week",
        "mobile_transpose": True,
    },
    {
        "id": "weekend-site-visits", "title": "Weekend Site Visits",
        "description": "Site visits scheduled for the upcoming Saturday and Sunday, by salesperson.",
        "cadence": "Weekly (Fri)", "admin_only": False, "owner_filterable": True,
        "caveat": None, "build": _build_weekend_site_visits, "date_default": "live",
        "template": "report_view_charts.html",
    },
    {
        "id": "monday-weekly-report", "title": "Monday Weekly Report",
        "description": "Leads, site visits, and bookings for the most recently completed Monday-Sunday week, company-wide and per salesperson.",
        "cadence": "Weekly (Mon)", "admin_only": False, "owner_filterable": True,
        "caveat": (
            "Site Visits Scheduled/Conducted use a different definition than the "
            "other Weekly Reports (direct site_visits dates, not activity_log) — "
            "numbers may not match those reports for the same week."
        ),
        "build": _build_monday_weekly_report, "date_default": "last_week",
        "template": "report_view_charts.html",
    },
    {
        "id": "lead-stage-analysis", "title": "Lead Stage Analysis",
        "description": "Lead counts by stage — broken down by owner, project, and campaign — plus site visits by campaign, for the selected period.",
        "cadence": "Monthly", "admin_only": False, "owner_filterable": True,
        "caveat": (
            "By-campaign and site-visits-by-campaign views are dominated by \"Unknown/Manual\" until campaign "
            "data is backfilled — see this report's Campaign data note."
        ),
        "build": _build_lead_stage_analysis, "date_default": "this_month",
        "template": "report_view_charts.html",
    },
    {
        "id": "campaign-lead-volume", "title": "Campaign Lead Volume",
        "description": "Which campaigns brought in the most leads during the selected period.",
        "cadence": "Marketing", "admin_only": True, "owner_filterable": False,
        "caveat": "leads.campaign is blank/NULL on the large majority of leads today — see this report's Campaign data note.",
        "build": _build_campaign_lead_volume, "date_default": "last_30_days",
        "template": "report_view_charts.html",
    },
    {
        "id": "campaign-stage-distribution", "title": "Campaign Stage Distribution",
        "description": "Where each campaign's leads end up in the pipeline, for the selected period.",
        "cadence": "Marketing", "admin_only": True, "owner_filterable": False,
        "caveat": "leads.campaign is blank/NULL on the large majority of leads today — see this report's Campaign data note.",
        "build": _build_campaign_stage_distribution, "date_default": "last_30_days",
        "template": "report_view_charts.html",
    },
    {
        "id": "campaign-conversion-rate", "title": "Campaign Conversion Rate",
        "description": "% of each campaign's leads reaching Site Visited+, for the selected period.",
        "cadence": "Marketing", "admin_only": True, "owner_filterable": False,
        "caveat": "leads.campaign is blank/NULL on the large majority of leads today — see this report's Campaign data note.",
        "build": _build_campaign_conversion_rate, "date_default": "last_30_days",
        "template": "report_view_charts.html",
    },
    {
        "id": "campaign-lost-reasons", "title": "Campaign Lost Reason Breakdown",
        "description": "Why leads from each campaign died, for the selected period.",
        "cadence": "Marketing", "admin_only": True, "owner_filterable": False,
        "caveat": "leads.campaign is blank/NULL on the large majority of leads today — see this report's Campaign data note.",
        "build": _build_campaign_lost_reasons, "date_default": "last_30_days",
    },
    {
        "id": "campaign-response-time", "title": "Campaign First-Response Time",
        "description": "Median time to first response, per campaign, for the selected period.",
        "cadence": "Marketing", "admin_only": True, "owner_filterable": False,
        "caveat": "leads.campaign is blank/NULL on the large majority of leads today — see this report's Campaign data note.",
        "build": _build_campaign_response_time, "date_default": "last_30_days",
        "template": "report_view_charts.html",
    },
    {
        "id": "campaign-site-visits", "title": "Site Visits by Campaign",
        "description": "Site visits per campaign, for the selected period. Standalone version of Lead Stage Analysis's view D.",
        "cadence": "Marketing", "admin_only": True, "owner_filterable": False,
        "caveat": "leads.campaign is blank/NULL on the large majority of leads today — see this report's Campaign data note.",
        "build": _build_campaign_site_visits, "date_default": "last_30_days",
        "template": "report_view_charts.html",
    },
]

REPORTS_BY_ID = {r["id"]: r for r in REPORTS}

# ─────────────────────────────────────────────────────────────
# CATEGORIES — config-not-code (Requirement 2). Landing page renders
# from this, in this order; a report id not listed here still works by
# direct URL, it just won't appear on /reports.
# ─────────────────────────────────────────────────────────────

CATEGORIES = [
    {"name": "Daily Operations", "report_ids": [
        "salesperson-scorecard", "pipeline-funnel", "owner-workload", "score-distribution",
    ]},
    {"name": "Weekly Reports", "report_ids": [
        "weekly-leads-received", "weekly-site-visits-conducted", "weekly-site-visits-scheduled",
        "weekend-site-visits", "monday-weekly-report",
    ]},
    {"name": "Pipeline Analysis", "report_ids": [
        "conversion-funnel", "project-pipeline", "lead-stage-analysis", "hit-rate", "first-response",
    ]},
    {"name": "Lead Quality & Retention", "report_ids": [
        "lost-reason", "reengagement",
    ]},
    {"name": "Marketing Performance", "report_ids": [
        "source-performance", "campaign-lead-volume", "campaign-stage-distribution",
        "campaign-conversion-rate", "campaign-lost-reasons", "campaign-response-time", "campaign-site-visits",
    ]},
    {"name": "Activity Log", "report_ids": [
        "call-activity",
    ]},
]


def visible_categories(current_user):
    """
    (v1.1) Landing-page structure — replaces v1.0's flat visible_reports().
    Same "admin_only reports simply absent, not greyed-out" rule as
    before, applied per-category; a category with zero visible reports
    for this login is omitted entirely rather than shown empty.
    """
    oversight = cls_db.can_view_all_leads(current_user["role"])
    result = []
    for cat in CATEGORIES:
        reports = [REPORTS_BY_ID[rid] for rid in cat["report_ids"] if oversight or not REPORTS_BY_ID[rid]["admin_only"]]
        if reports:
            result.append({"name": cat["name"], "reports": reports})
    return result


def build_report(report_id, current_user, date_from=None, date_to=None, **kwargs):
    """
    Runs one report's builder and returns the full renderable dict:
    {id, title, description, cadence, caveat, date_default, is_live,
    template, date_from, date_to, columns, rows, chart, views}.
    Raises KeyError for an unknown report_id — the route treats that
    as a 404, never a 500.
    """
    meta = REPORTS_BY_ID[report_id]
    table = meta["build"](current_user, date_from=date_from, date_to=date_to, **kwargs)
    columns = table.get("columns", [])
    rows = table.get("rows", [])
    result = {
        "id": meta["id"], "title": meta["title"], "description": meta["description"],
        "cadence": meta["cadence"], "caveat": meta["caveat"],
        "date_default": meta.get("date_default", "this_month"),
        "is_live": meta.get("date_default") == "live",
        "template": meta.get("template", "report_view.html"),
        "date_from": date_from, "date_to": date_to,
        # v1.2 — Task 1: which preset (if any) the resolved date_from/
        # date_to matches, so the picker dropdown can pre-select it.
        # None for "live" reports (date_from is None, not ""), which
        # never show the picker at all.
        "date_preset": _detect_active_report_preset(date_from, date_to) if date_from is not None else None,
        "date_preset_order": REPORT_DATE_PRESET_ORDER,
        "date_preset_labels": REPORT_DATE_PRESET_LABELS,
        "columns": columns, "rows": rows,
        "chart": table.get("chart"), "views": table.get("views"),
    }
    # v1.2 — Task 2: mobile transpose, additive alongside columns/rows
    # (export_to_excel() keeps reading columns/rows only, untouched).
    # The row-identity column is always this report's FIRST column for
    # all 5 mobile_transpose reports (verified against each builder).
    if meta.get("mobile_transpose") and columns and rows:
        label_key = columns[0][0]
        display_columns, display_rows = transpose_table(columns, rows, label_key)
        result["display_columns"] = display_columns
        result["display_rows"] = display_rows
    return result


# ─────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────

def _report_sheets(report):
    """
    (v1.1) Normalizes a built report into a list of {title, columns,
    rows} — one sheet per view for multi-view reports (Lead Stage
    Analysis), or a single sheet for every other report, unchanged from
    v1.0's shape.
    """
    if report.get("views"):
        return [{"title": v["title"], "columns": v["columns"], "rows": v["rows"]} for v in report["views"]]
    return [{"title": report["title"], "columns": report["columns"], "rows": report["rows"]}]


def export_to_excel(report):
    """
    report : the dict build_report() returns. Writes one worksheet per
    _report_sheets() entry, header row bolded, column widths auto-sized
    to content (capped so one long free-text cell can't blow up the
    whole sheet). Every sheet's second row states the report's title and
    the exact date range it was generated for (or omits the range for a
    "live" report), so an exported file is self-describing even once
    detached from the app.

    Returns a BytesIO positioned at 0, ready for Flask's send_file().
    Raises RuntimeError if openpyxl isn't installed — the route
    catches this and shows a clear message rather than a raw 500.
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError(
            "openpyxl is not installed. Run: python -m pip install openpyxl"
        )

    wb = Workbook()
    first = True
    range_note = f"{report['date_from']} to {report['date_to']}" if report.get("date_from") else None

    for sheet_data in _report_sheets(report):
        ws = wb.active if first else wb.create_sheet()
        first = False
        # Excel sheet titles: max 31 chars, and none of \ / ? * [ ] allowed —
        # several report/view titles ("Lost/Unqualified...") contain a slash.
        safe_title = sheet_data["title"]
        for ch in r'\/?*[]':
            safe_title = safe_title.replace(ch, "-")
        ws.title = safe_title[:31]

        ws.append([report["title"]])
        ws["A1"].font = Font(bold=True, size=13)
        generated_line = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        if range_note:
            generated_line += f" — {range_note}"
        ws.append([generated_line])
        if report.get("caveat"):
            ws.append([report["caveat"]])
        ws.append([])

        header_row_idx = ws.max_row + 1
        labels = [label for _, label in sheet_data["columns"]]
        ws.append(labels)
        for cell in ws[header_row_idx]:
            cell.font = Font(bold=True)

        for row in sheet_data["rows"]:
            ws.append([row.get(key, "") for key, _ in sheet_data["columns"]])

        for i, (key, label) in enumerate(sheet_data["columns"], start=1):
            col_letter = ws.cell(row=header_row_idx, column=i).column_letter
            max_len = max(
                [len(label)] + [len(str(row.get(key, ""))) for row in sheet_data["rows"]]
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
